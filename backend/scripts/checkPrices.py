#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Verifică prețurile produselor PilotOn (navigatii-gps) față de grila de prețuri
bazată pe varianta hardware + prețul ramei din Excel.

Doar citește și raportează — nu modifică nimic.

Utilizare:
    python3 checkPrices.py                 # fetch live din API
    python3 checkPrices.py --cached        # refolosește cache-ul local de produse
    python3 checkPrices.py --excel /path/to/lista.xlsx

Ieșire: price-check-report.json + sumar în consolă.
"""

import argparse
import json
import os
import re
import sys
import time
import unicodedata
import urllib.request
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
API = os.environ.get('PILOTON_API', 'https://api.navi.piloton.ro')
DEFAULT_EXCEL = os.path.expanduser('~/Downloads/LISTA APARATE +RAME ACTUALIZATA 11.2025 2.xlsx')
CACHE_FILE = os.path.join(HERE, 'price-check-products-cache.json')
REPORT_FILE = os.path.join(HERE, 'price-check-report.json')

SKIP_SHEETS = {'Rezumat exportare', 'APARATE PILOTON', 'APARATE JUNSUN ',
               'ADAPTOARE 2DIN VW, SKODA , SEAT'}

# familie -> (sub 100, 100-200 inclusiv, peste 200)
RAMA_FAMILIES = {
    '2GB RK':        (849, 899, 949),
    '4GB RK':        (999, 1049, 1099),
    '4GB XT':        (1399, 1449, 1499),
    '6GB XT':        (1599, 1649, 1699),
    '4GB 8667':      (1599, 1649, 1699),
    '8GB 8667':      (2299, 2349, 2399),
    '4GB 8667 2K':   (1999, 2049, 2099),
    '8GB 8667 2K':   (2799, 2849, 2899),
}
PRICE_12GB = 3299  # fix, indiferent de ramă

# Grila OEM: (grup, diagonală) -> {ram: preț}
OEM_BMW_STD = {'10.25': {4: 2475, 8: 2795}, '12.3': {4: 2795, 8: 3115}, '12.9': {4: 3295, 8: 3615}}
OEM_BMW_EVO = {'10.25': {4: 2775, 8: 3095}, '12.3': {4: 3095, 8: 3415}, '12.9': {4: 3595, 8: 3915}}
OEM_MERCEDES = OEM_BMW_STD  # NTG 4.0/4.5/5.0 = aceeași grilă ca BMW CIC/CCC/NBT
OEM_AUDI = {'8.8': {4: 2475, 8: 2990}, '10.25': {4: 2690, 8: 3115}}
OEM_AUDI_A6C7 = {'9': {4: 3130, 8: 3450}, '12.3': {4: 3130, 8: 3450}}
OEM_AUDI_Q_123 = {8: 3450}  # Q3/Q5 12.3"

BMW_EVO_MARKERS = {'EVO', 'NBT EVO', 'NOS', 'NJS', 'NSJ'}  # NSJ = NJS scris invers în nume
BMW_STD_MARKERS = {'CIC', 'CCC', 'NBT'}

BRANDS = ['ALFA ROMEO', 'ALFA', 'MERCEDES BENZ', 'MERCEDES', 'LAND ROVER', 'VW', 'VOLKSWAGEN',
          'SEAT', 'SKODA', 'AUDI', 'BMW', 'RENAULT', 'DACIA', 'MAZDA', 'SUZUKI', 'NISSAN',
          'PEUGEOT', 'MITSUBISHI', 'HYUNDAI', 'LANCIA', 'VOLVO', 'HONDA', 'KIA', 'OPEL',
          'SMART', 'PORSCHE', 'PORCHE', 'CHRYSLER', 'IVECO', 'JEEP', 'FORD', 'FIAT',
          'CITROEN', 'SUBARU', 'CHEVROLET', 'LEXUS', 'TOYOTA', 'TOYORA', 'DODGE',
          'ISUZU', 'SSANGYONG', 'JAGUAR', 'MINI', 'SAAB', 'DAEWOO', 'INFINITI', 'CADILLAC',
          'UNIVERSAL', 'UNIVERSALA', '2DIN UNIVERSALA']
BRAND_CANON = {'ALFA': 'ALFA ROMEO', 'MERCEDES': 'MERCEDES BENZ', 'VOLKSWAGEN': 'VW',
               'PORCHE': 'PORSCHE', 'TOYORA': 'TOYOTA',
               'UNIVERSALA': 'UNIVERSAL', '2DIN UNIVERSALA': 'UNIVERSAL'}

# corecturi de typo-uri / unificări aplicate DUPĂ uppercase, pe text întreg
TEXT_FIXES = [
    (r'\bBETTLE\b', 'BEETLE'), (r'\bTOYORA\b', 'TOYOTA'),
    (r'\bFOR\s+TOW\b', 'FORTWO'), (r'\bFOR\s+TWO\b', 'FORTWO'),
    (r'\bRAV\s*4\b', 'RAV4'), (r'\bT\s+CROSS\b', 'TCROSS'), (r'\bT\s+ROC\b', 'TROC'),
    (r'\bACORD\b', 'ACCORD'), (r'\bCIVIV\b', 'CIVIC'), (r'\bJYMNI\b', 'JIMNY'),
    (r'\bSANRAFE\b', 'SANTAFE'), (r'\bSANTA\s+FE\b', 'SANTAFE'), (r'\bAUT0\b', 'AUTO'),
    (r'\bDAYLI\b', 'DAILY'), (r'\bJEPP\b', 'JEEP'), (r'\bLACER\b', 'LANCER'),
    (r'\bMITSUBSHI\b', 'MITSUBISHI'), (r'\bFIAR\b', 'FIAT'), (r'\bCERNATO\b', 'CERATO'),
    (r'\bKARENS\b', 'CARENS'), (r'\bTIDA\b', 'TIIDA'), (r'\bELYSSE\b', 'ELYSEE'),
    (r'\bFOREST\b', 'FORESTER'), (r'\bIS250\b', 'IS 250'), (r'\bMK(\d)\b', r'\1'),
    (r'\bBT\s+50\b', 'BT50'), (r'\bS\s+MAX\b', 'SMAX'), (r'\bC\s+MAX\b', 'CMAX'),
    (r'\bX\s+TRAIL\b', 'XTRAIL'), (r'\b500\s+X\b', '500X'), (r'\b500\s+L\b', '500L'),
]

# numerale romane folosite ca generație (Excel folosește cifre: ACORD 8, MONDEO 4)
ROMAN = {'II': '2', 'III': '3', 'IV': '4', 'VI': '6', 'VII': '7', 'VIII': '8', 'IX': '9'}

# litere din Excel <-> trim-uri din numele produselor
TRIM_LETTERS = {'S': 'SILVER', 'B': 'BLACK', 'G': 'GREY', 'W': 'WHITE'}
TRIM_WORDS = {'SILVER', 'BLACK', 'GREY', 'GRAY', 'WHITE', 'CLIMA', 'AC', 'GP', 'GT',
              'HIGH', 'LOW', 'WOOD', 'MANUAL', 'AUTO', 'BEJ', 'MAHON',
              'RAMA', 'SIMPLA', 'CU', 'BUTOANE', 'GROASA', 'SUBTIRE', 'MARE', 'MICA',
              'SUS', 'JOYSTICK', 'BOX', 'LVDS', 'DOORS', '1DIN', '2DIN'}


def log(msg):
    print(msg, flush=True)


def strip_diacritics(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def http_json(url, tries=5):
    last = None
    for i in range(tries):
        try:
            with urllib.request.urlopen(url, timeout=90) as r:
                return json.loads(r.read())
        except Exception as e:
            last = e
            time.sleep(2 * (i + 1))
    raise RuntimeError(f'GET {url} a eșuat după {tries} încercări: {last}')


# ---------------------------------------------------------------- Excel

def parse_excel(path, warnings):
    """Returnează lista de rânduri {brand, model, raw, size, rama, sheet}.

    Scanează fiecare sheet după TOATE tripletele de headere MODEL/SIZE/PRICE —
    unele sheet-uri au mai multe blocuri pe coloane diferite (SKODA are și
    rânduri FORD în col. A-C, SMART are blocul duplicat).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sheet in wb.sheetnames:
        if sheet in SKIP_SHEETS:
            continue
        ws = wb[sheet]
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
        # găsește coloanele de start ale blocurilor MODEL/SIZE/PRICE
        blocks = []  # (row_idx, col_idx)
        for ri, row in enumerate(grid):
            for ci in range(len(row) - 2):
                a = str(row[ci] or '').strip().upper()
                b = str(row[ci + 1] or '').strip().upper()
                c = str(row[ci + 2] or '').strip().upper()
                if a == 'MODEL' and b == 'SIZE' and c == 'PRICE':
                    blocks.append((ri, ci))
        if not blocks:
            # sheet fără rând de header (ex. MINI/ISUZU adăugate manual) — citesc direct col. A-C
            blocks = [(-1, 0)]
        for (hri, ci) in blocks:
            for row in grid[hri + 1:]:
                if ci + 2 >= len(row):
                    continue
                model, size, price = row[ci], row[ci + 1], row[ci + 2]
                if model is None and price is None:
                    continue
                if not isinstance(price, (int, float)) or not model:
                    if model or price:
                        warnings.append(f'Sheet "{sheet}": rând ignorat: {model!r} / {size!r} / {price!r}')
                    continue
                raw = re.sub(r'\s+', ' ', str(model)).strip()
                rows.append({'sheet': sheet, 'raw': raw, 'size': int(size) if size else None,
                             'rama': float(price)})
    # dubluri exacte cu prețuri diferite
    seen = defaultdict(set)
    for r in rows:
        seen[(r['raw'].upper(), r['size'])].add(r['rama'])
    for (raw, size), prices in seen.items():
        if len(prices) > 1:
            warnings.append(f'Excel: "{raw}" (size {size}) apare cu prețuri diferite: {sorted(prices)}')
    return rows


# ---------------------------------------------------------------- normalizare & potrivire

YEAR_RANGE = re.compile(r'\b(\d{2}|\d{4})\s*-\s*(\d{2}|\d{4})\b')
YEAR_PLUS = re.compile(r'\b(\d{2}|\d{4})\s*\+')
YEAR_DUPA = re.compile(r'\bDUPA\s+(\d{4})\b')
YEAR_SINGLE = re.compile(r'\b(19\d{2}|20\d{2})\b')


def to_year(t):
    y = int(t)
    if y < 100:
        y += 2000 if y < 50 else 1900
    return y


def extract_years(text):
    """Returnează (text_fara_ani, (start, end) sau None). end=None => deschis."""
    span = None

    def rng(m):
        nonlocal span
        a, b = to_year(m.group(1)), to_year(m.group(2))
        if a > b:
            a, b = b, a
        span = (a, b)
        return ' '

    def plus(m):
        nonlocal span
        span = (to_year(m.group(1)), None)
        return ' '

    text = YEAR_DUPA.sub(plus, text)
    text = YEAR_RANGE.sub(rng, text)
    text = YEAR_PLUS.sub(plus, text)
    if span is None:
        m = YEAR_SINGLE.search(text)
        if m:
            span = (int(m.group(1)), int(m.group(1)))
            text = text[:m.start()] + ' ' + text[m.end():]
    return text, span


def year_sim(a, b):
    """Similaritatea intervalelor de ani: 0 = fără suprapunere, 1 = incluziune."""
    if a is None or b is None:
        return None  # necunoscut
    cap = 2026
    a0, a1 = a[0], a[1] if a[1] is not None else cap
    b0, b1 = b[0], b[1] if b[1] is not None else cap
    ov = min(a1, b1) - max(a0, b0) + 1
    if ov <= 0:
        return 0.0
    return ov / (min(a1 - a0, b1 - b0) + 1)


def normalize(text):
    t = strip_diacritics(str(text)).upper()
    t = t.replace('/', ' ').replace(',', ' ').replace('*', ' ')
    # unește C-MAX/CR-V/S-MAX/C-HR/T-ROC/CX-5 etc. (nu și intervalele de ani cifră-cifră)
    t = re.sub(r'(?<=[A-Z])\s*-\s*(?=[A-Z0-9])', '', t)
    for pat, rep in TEXT_FIXES:
        t = re.sub(pat, rep, t)
    t = re.sub(r'\s+', ' ', t).strip()
    return t


def detect_brand(norm_text):
    for b in BRANDS:
        if norm_text.startswith(b + ' ') or norm_text == b:
            return BRAND_CANON.get(b, b)
    return None


def find_all_brands(norm_text):
    """Toate mărcile menționate oriunde în text (rânduri combinate gen
    "TOYOTA AYGO / CITROEN C1")."""
    found = set()
    for b in BRANDS:
        if re.search(r'\b' + re.escape(b) + r'\b', norm_text):
            found.add(BRAND_CANON.get(b, b))
    return found


def tokenize(norm_text):
    t, span = extract_years(norm_text)
    # scoate toate numele de mărci din tokenuri (rămâne doar modelul)
    for b in BRANDS:
        t = re.sub(r'\b' + re.escape(b) + r'\b', ' ', t)
    toks = set()
    for w in re.split(r'[^A-Z0-9.+]+', t):
        w = w.strip('.')
        if not w:
            continue
        w = TRIM_LETTERS.get(w, w)
        w = ROMAN.get(w, w)
        toks.add(w)
    return toks, span


# BMW în Excel e pe cod de șasiu, produsele pe „Seria N ani" — echivalări cunoscute
BMW_CHASSIS = [
    ({'SERIA', '3'}, (1997, 2007), 'E46'),
    ({'SERIA', '3'}, (2004, 2013), 'E90'),
    ({'SERIA', '5'}, (1995, 2004), 'E39'),
    ({'X5'}, (1999, 2007), 'E53'),
]


def augment_tokens(brand, tokens, years):
    if brand == 'BMW' and years:
        best = max(((year_sim(years, span) or 0.0, code)
                    for req, span, code in BMW_CHASSIS if req <= tokens),
                   default=(0.0, None))
        if best[0] > 0:
            tokens = tokens | {best[1]}
    return tokens


class ExcelEntry:
    def __init__(self, row):
        self.raw = row['raw']
        self.sheet = row['sheet']
        self.size = row['size']
        self.rama = row['rama']
        norm = normalize(row['raw'])
        self.brands = find_all_brands(norm)
        if not self.brands:
            sheet_brand = detect_brand(normalize(row['sheet']))
            self.brands = {sheet_brand} if sheet_brand else set()
        self.tokens, self.years = tokenize(norm)

    def label(self):
        return f'{self.raw} (size {self.size}, rama {self.rama:g}) [{self.sheet}]'


def match_model(model_str, size, entries_by_brand):
    """Potrivește un șir de model de produs cu un rând din Excel.

    Returnează (entry, note) sau (None, motiv).
    """
    norm = normalize(model_str)
    brand = detect_brand(norm)
    if brand is None:
        return None, f'marcă necunoscută în "{model_str}"'
    tokens, years = tokenize(norm)
    tokens = augment_tokens(brand, tokens, years)
    candidates = entries_by_brand.get(brand, [])
    if not candidates:
        return None, f'nicio intrare în Excel pentru marca {brand}'

    chassis = re.compile(r'^[A-Z]{1,2}\d{2,3}$')  # W203, E90, J10, IX35, CX5…
    scored = []
    for e in candidates:
        ysim = year_sim(years, e.years)
        core_excel = e.tokens - TRIM_WORDS
        core_prod = tokens - TRIM_WORDS
        core_inter = core_prod & core_excel
        if core_excel and core_prod and not core_inter:
            continue
        # anii pot contrazice doar dacă nu există un cod de șasiu comun (Excelul
        # are uneori alți ani decât numele produsului pentru același W/E-cod)
        if ysim == 0.0:
            if not any(chassis.match(t) for t in core_inter):
                continue
            ysim = 0.1
        # potrivirea doar pe cifra de generație (ex. „KUGA 2" vs „Galaxy II") nu contează
        if core_inter and all(t.isdigit() for t in core_inter) \
                and any(not t.isdigit() for t in core_excel):
            continue
        score = 0.0
        if core_excel:
            score += 2.0 * len(core_inter) / len(core_excel)
        elif not core_prod:
            score += 2.0  # ambele doar marcă (ex. UNIVERSAL)
        else:
            score += 1.0  # Excel are doar marca, produsul are model
        if core_prod:
            score += 1.0 * len(core_inter) / len(core_prod)
        # trim: bonus per potrivire, penalty per trim din Excel neacoperit de produs
        # (astfel „YARIS S / B GP" pierde în fața „YARIS S / B" când produsul n-are GP)
        trim_prod = tokens & TRIM_WORDS
        trim_excel = e.tokens & TRIM_WORDS
        score += 0.5 * len(trim_prod & trim_excel) - 0.15 * len(trim_excel - trim_prod)
        score += 0.25 if ysim is None else ysim
        # preferă mărimea corectă; rând pe cealaltă mărime rămâne fallback
        size_match = (e.size == size)
        scored.append((score, size_match, e))

    if not scored:
        return None, 'niciun rând compatibil (model/ani) în Excel'
    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    best_score = scored[0][0]
    if best_score < 1.8:
        return None, f'scor prea mic ({best_score:.2f}) — cel mai apropiat: {scored[0][2].label()}'
    top = [s for s in scored if s[0] >= best_score - 0.05]
    with_size = [s for s in top if s[1]]
    pool = with_size or top
    ramas = {s[2].rama for s in pool}
    if len(ramas) > 1 and years:
        # departajare: rândul al cărui an de început e cel mai aproape de al produsului
        dists = sorted(((abs(s[2].years[0] - years[0]) if s[2].years else 99, s)
                        for s in pool), key=lambda x: x[0])
        if dists[0][0] < dists[1][0]:
            pool = [dists[0][1]]
            ramas = {pool[0][2].rama}
    if len(ramas) > 1:
        opts = '; '.join(s[2].label() for s in pool[:4])
        return None, f'ambiguu — {len(pool)} rânduri cu rame diferite: {opts}'
    entry = pool[0][2]
    note = None
    if entry.size != size:
        note = f'folosit rândul de size {entry.size} (nu există size {size})'
    return entry, note


# ---------------------------------------------------------------- clasificare produse

def get_proc(p):
    rs = p.get('romanianSpecs') or {}
    hw = rs.get('hardware') or {}
    return (hw.get('modelProcesor') or '').upper()


def classify(p, warnings):
    """Returnează dict cu: kind ('rama'|'fixed'|'unchecked'), family/expected/reason, size."""
    name = p['name']
    proc = get_proc(p)

    m = re.search(r'(\d+)\s*GB\s+(\d+)\s*GB', name, re.I)
    ram = int(m.group(1)) if m else None
    is2k = bool(re.search(r'\b2K\b', name, re.I))
    tesla = bool(re.search(r'tesla', name, re.I))
    im = re.search(r'(\d+(?:\.\d+)?)\s*inch', name, re.I)
    inch = im.group(1) if im else None
    cores8 = bool(re.search(r'8\s*Core|Octa', name, re.I))

    oem_markers = set(re.sub(r'\s+', ' ', mm.upper()) for mm in re.findall(
        r'\b(CIC|CCC|NBT\s*EVO|NBT|EVO|NOS|NJS|NSJ|NTG|MMI3G|MMI2G|MMI|RMC|MHI2|HN\+?R)\b',
        name, re.I))
    oem_inch = inch in {'8.8', '8.9', '10.25', '12.3', '12.9'} if inch else False

    if tesla:
        return {'kind': 'unchecked', 'reason': 'Tesla 9.7" — fără regulă'}
    if inch in {'7', '8'}:
        return {'kind': 'unchecked', 'reason': f'{inch} inch — fără regulă'}
    if re.search(r'\b3K\b', name, re.I):
        return {'kind': 'unchecked', 'reason': 'ecran 3K — fără regulă'}

    # --- ecrane OEM (grilă fixă, fără ramă)
    if oem_inch or oem_markers:
        return classify_oem(name, ram, inch, oem_markers)

    if ram == 12:
        return {'kind': 'fixed', 'family': '12GB 2K', 'expected': PRICE_12GB}

    if ram is None:
        return {'kind': 'unchecked', 'reason': 'fără RAM/ROM în nume'}

    rom = int(m.group(2))
    is8667 = '8667' in proc

    family = None
    if (ram, rom) == (2, 32) and not cores8:
        family = '2GB RK'
    elif (ram, rom) == (4, 64) and not cores8:
        family = '4GB RK'
    elif (ram, rom) == (4, 64) and cores8:
        if is2k:
            family = '4GB 8667 2K'
        else:
            family = '4GB 8667' if is8667 else '4GB XT'
    elif (ram, rom) == (6, 128):
        family = '6GB XT' if not is2k else None
    elif (ram, rom) == (8, 256):
        family = '8GB 8667 2K' if is2k else '8GB 8667'

    if family is None:
        return {'kind': 'unchecked', 'reason': f'{ram}GB+{rom}GB{" 2K" if is2k else ""} — fără regulă'}

    if not proc and family in ('4GB XT', '4GB 8667'):
        warnings.append(f'{p["slug"]}: fără modelProcesor în specs — presupus XT')

    # mărimea pentru alegerea ramei
    if inch is not None and inch not in ('9', '10', '10.1'):
        return {'kind': 'unchecked', 'reason': f'{inch} inch — fără regulă'}
    if inch in ('9',):
        size = 9
    elif inch in ('10', '10.1'):
        size = 10
    elif is2k:
        size = 10 if re.search(r'2K\s*10', name, re.I) else 9
    else:
        size = 9
        warnings.append(f'{p["slug"]}: fără inch în nume — presupus 9"')
    return {'kind': 'rama', 'family': family, 'size': size}


def classify_oem(name, ram, inch, markers):
    n = name.upper()
    if ram not in (4, 8):
        # 8+128 pe ecrane OEM le tratăm ca 8GB; altfel n-avem coloană
        if ram == 6 or ram is None:
            return {'kind': 'unchecked', 'reason': f'OEM cu {ram}GB — fără regulă'}
    ramcol = 8 if ram == 8 else 4

    bucket = None
    if inch in ('8.8', '8.9'):
        bucket = '8.8'
    elif inch == '10.25':
        bucket = '10.25'
    elif inch == '12.3':
        bucket = '12.3'
    elif inch == '12.9':
        bucket = '12.9'

    if 'BMW' in n:
        grid = OEM_BMW_EVO if (markers & BMW_EVO_MARKERS) else OEM_BMW_STD
        gname = 'BMW EVO/NoS/NJS' if grid is OEM_BMW_EVO else 'BMW CIC/CCC/NBT'
        b = '10.25' if bucket == '8.8' else bucket  # presupunere: BMW 8.9" = treapta 10.25
        if b in grid:
            return {'kind': 'fixed', 'family': f'{gname} {b}" {ramcol}GB', 'expected': grid[b][ramcol]}
        return {'kind': 'unchecked', 'reason': f'BMW OEM fără diagonală cunoscută ({inch}")'}

    if 'MERCEDES' in n:
        if bucket and bucket != '8.8':
            return {'kind': 'fixed', 'family': f'Mercedes NTG {bucket}" {ramcol}GB',
                    'expected': OEM_MERCEDES[bucket][ramcol]}
        return {'kind': 'unchecked', 'reason': f'Mercedes OEM fără diagonală cunoscută ({inch}")'}

    if 'AUDI' in n:
        if re.search(r'\bA6\s*C7\b', n):
            b = '12.3' if bucket == '12.3' else '9'
            return {'kind': 'fixed', 'family': f'Audi A6 C7 {b}" {ramcol}GB',
                    'expected': OEM_AUDI_A6C7[b][ramcol]}
        if re.search(r'\bQ[35]\b', n) and bucket == '12.3':
            if ramcol == 8:
                return {'kind': 'fixed', 'family': 'Audi Q3/Q5 12.3" 8GB', 'expected': OEM_AUDI_Q_123[8]}
            return {'kind': 'unchecked', 'reason': 'Audi Q3/Q5 12.3" 4GB — fără regulă'}
        if bucket in OEM_AUDI:
            return {'kind': 'fixed', 'family': f'Audi {bucket}" {ramcol}GB',
                    'expected': OEM_AUDI[bucket][ramcol]}
        return {'kind': 'unchecked', 'reason': f'Audi OEM fără diagonală cunoscută ({inch}")'}

    return {'kind': 'unchecked', 'reason': f'marker OEM ({", ".join(sorted(markers)) or inch}") la marcă fără grilă OEM'}


MODEL_CUT = re.compile(r'\s+(?:\b2K\b|\d+(?:\.\d+)?\s*inch\b|\d+\s*GB\b)', re.I)


def model_part(name):
    rest = re.sub(r'^Navigatie PilotOn\s+', '', name, flags=re.I)
    m = MODEL_CUT.search(rest)
    return (rest[:m.start()] if m else rest).strip()


def rama_tier(rama):
    if rama < 100:
        return 0
    if rama <= 200:
        return 1
    return 2


# ---------------------------------------------------------------- main

def fetch_products(use_cache):
    if use_cache and os.path.exists(CACHE_FILE):
        log(f'Folosesc cache-ul {CACHE_FILE}')
        with open(CACHE_FILE) as f:
            return json.load(f)
    products, page = [], 1
    while True:
        # sortBy=_id e obligatoriu: sortarea implicită (createdAt) dă 500 pe colecția mare
        d = http_json(f'{API}/api/products?limit=200&page={page}&sortBy=_id&sortOrder=asc')
        for p in d.get('products', []):
            products.append({
                'name': p['name'], 'slug': p.get('slug'), 'price': p.get('price'),
                'brand': p.get('brand'), 'category': p.get('category'), 'status': p.get('status'),
                'romanianSpecs': {'hardware': (p.get('romanianSpecs') or {}).get('hardware')},
            })
        pg = d.get('pagination', {})
        log(f'  pagina {page}/{pg.get("totalPages", "?")} — {len(products)} produse')
        if not pg.get('hasNextPage'):
            break
        page += 1
    with open(CACHE_FILE, 'w') as f:
        json.dump(products, f, ensure_ascii=False)
    return products


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', default=DEFAULT_EXCEL)
    ap.add_argument('--cached', action='store_true', help='refolosește cache-ul local de produse')
    args = ap.parse_args()

    warnings = []
    log(f'Citesc Excel: {args.excel}')
    excel_rows = parse_excel(args.excel, warnings)
    entries = [ExcelEntry(r) for r in excel_rows]
    entries_by_brand = defaultdict(list)
    for e in entries:
        for b in e.brands:
            entries_by_brand[b].append(e)
    log(f'  {len(entries)} rânduri de ramă în {len(entries_by_brand)} mărci')

    log('Aduc produsele din API…')
    products = fetch_products(args.cached)
    log(f'  {len(products)} produse în total')

    in_scope = [p for p in products
                if p.get('category') == 'navigatii-gps'
                and re.match(r'Navigatie PilotOn\b', p.get('name') or '', re.I)]
    out_of_scope = len(products) - len(in_scope)

    mismatches, ok, unmatched, unchecked = [], [], [], []
    used_excel = set()
    match_cache = {}

    for p in in_scope:
        c = classify(p, warnings)
        if c['kind'] == 'unchecked':
            unchecked.append({'slug': p['slug'], 'name': p['name'], 'price': p['price'],
                              'reason': c['reason']})
            continue

        if c['kind'] == 'fixed':
            expected, detail = c['expected'], {'family': c['family']}
        else:
            ms = model_part(p['name'])
            key = (ms, c['size'])
            if key not in match_cache:
                match_cache[key] = match_model(ms, c['size'], entries_by_brand)
            entry, note = match_cache[key]
            if entry is None:
                unmatched.append({'slug': p['slug'], 'name': p['name'], 'price': p['price'],
                                  'family': c['family'], 'model': ms, 'reason': note})
                continue
            used_excel.add(id(entry))
            expected = RAMA_FAMILIES[c['family']][rama_tier(entry.rama)]
            detail = {'family': c['family'], 'excel': entry.label(), 'rama': entry.rama}
            if note:
                detail['note'] = note

        if p['price'] == expected:
            ok.append(p['slug'])
        else:
            mismatches.append({'slug': p['slug'], 'name': p['name'],
                               'price': p['price'], 'expected': expected,
                               'diff': round((p['price'] or 0) - expected, 2), **detail})

    unused_rows = [e.label() for e in entries if id(e) not in used_excel]

    report = {
        'generated': time.strftime('%Y-%m-%d %H:%M:%S'),
        'excel': args.excel,
        'summary': {
            'total_produse': len(products), 'in_scope': len(in_scope),
            'out_of_scope': out_of_scope, 'ok': len(ok), 'mismatches': len(mismatches),
            'unmatched': len(unmatched), 'unchecked': len(unchecked),
            'excel_rows': len(entries), 'unused_excel_rows': len(unused_rows),
        },
        'assumptions': [
            'Trepte ramă: sub 100 = mic, 100-200 inclusiv = mediu, peste 200 = mare',
            '12GB 2K = 3299 fix, indiferent de ramă',
            'BMW 8.8/8.9" = treapta 10.25 din grila OEM',
            'Audi 8.8 din reguli = "8.9 inch" din numele produselor',
            '2K fără inch în nume = 9", "2K 10" = 10"',
            '8+128GB pe ecrane OEM tratat ca 8GB',
        ],
        'mismatches': sorted(mismatches, key=lambda x: (x['family'], x['name'])),
        'unmatched': sorted(unmatched, key=lambda x: x['name']),
        'unchecked': sorted(unchecked, key=lambda x: (x['reason'], x['name'])),
        'unused_excel_rows': sorted(unused_rows),
        'warnings': warnings,
        'ok_slugs': ok,
    }
    with open(REPORT_FILE, 'w') as f:
        json.dump(report, f, ensure_ascii=False, indent=1)

    # ------- sumar consolă
    log('\n================ SUMAR ================')
    for k, v in report['summary'].items():
        log(f'  {k}: {v}')
    by_family = defaultdict(lambda: [0, 0])
    for mm in mismatches:
        by_family[mm['family']][0] += 1
    for s in ok:
        pass
    log('\nGreșite pe familii:')
    fam_diffs = defaultdict(list)
    for mm in mismatches:
        fam_diffs[mm['family']].append(mm['diff'])
    for fam in sorted(fam_diffs):
        diffs = fam_diffs[fam]
        log(f'  {fam:28s} {len(diffs):4d} produse  (dif. medie {sum(diffs)/len(diffs):+.0f} lei)')
    log('\nPrimele 15 nepotriviri de preț:')
    for mm in mismatches[:15]:
        log(f'  {mm["price"]:>7} → {mm["expected"]:>5}  {mm["name"]}')
    if unmatched:
        log(f'\nPrimele 10 fără rând în Excel (din {len(unmatched)}):')
        for u in unmatched[:10]:
            log(f'  [{u["family"]}] {u["model"]} — {u["reason"]}')
    log(f'\nRaport complet: {REPORT_FILE}')


if __name__ == '__main__':
    main()
