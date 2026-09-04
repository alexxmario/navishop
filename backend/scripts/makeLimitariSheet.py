#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Face tabelul de decizie pentru Alex din limitari-live-harvest.json:
textul lor verbatim + linkul + care dintre modelele NOASTRE ar fi atinse.

  python3 backend/scripts/makeLimitariSheet.py            # scrie limitari-decizie.csv
  python3 backend/scripts/makeLimitariSheet.py --xlsx     # si .xlsx daca ai openpyxl

CSV-ul e UTF-8 cu BOM, ca sa se deschida corect direct in Numbers si in Excel.
Coloana „Se aplica la noi?" e goala — o completeaza Alex.
"""
import json, os, csv, sys, re, argparse, importlib.util

HERE = os.path.dirname(os.path.abspath(__file__))
HARVEST = os.path.join(HERE, 'limitari-live-harvest.json')
OUR_MODELS = os.path.join(HERE, 'our-car-models.json')
CSV_OUT = os.path.join(HERE, 'limitari-decizie.csv')
XLSX_OUT = os.path.join(HERE, 'limitari-decizie.xlsx')

_spec = importlib.util.spec_from_file_location('mcm', os.path.join(HERE, 'matchCarModels.py'))
mcm = importlib.util.module_from_spec(_spec)
_argv, sys.argv = sys.argv, [sys.argv[0]]
_spec.loader.exec_module(mcm)
sys.argv = _argv

# clasificare grosiera, ca sa se vada dintr-o privire ce fel de limitare e
def clasifica(txt):
    t = mcm.deacc(txt or '')
    if 'nu se poate monta' in t and 'navigatie originala' in t:
        return 'incompatibilitate cu navigatia de fabrica (PROBABIL limitarea LOR, nu a masinii)'
    if 'cablu suplimentar' in t or 'adaptor suplimentar' in t:
        return 'cere cablu/adaptor separat'
    if 'butoane' in t and 'reloca' in t:
        return 'butoane relocate pe rama (tine de bord)'
    if any(w in t for w in ('bose', 'harman', 'jbl', 'amplificator')):
        return 'sistem audio amplificat din fabrica'
    if 'fibra optica' in t or ' most ' in t:
        return 'magistrala pe fibra optica'
    if 'clima' in t:
        return 'afisaj/comenzi clima'
    return 'altceva — de citit'


def parse_their_slug(slug):
    """'bmw-seria-1-e81-e87-2003-2013' -> ('bmw seria 1 e81 e87', '2003-2013')"""
    return mcm.parse_group_key(slug)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', action='store_true', help='scrie si .xlsx (necesita openpyxl)')
    args = ap.parse_args()

    harvest = json.load(open(HARVEST, encoding='utf-8'))['results']
    ours = json.load(open(OUR_MODELS, encoding='utf-8')) if os.path.exists(OUR_MODELS) else []
    if not ours:
        print(f'ATENTIE: lipseste {OUR_MODELS} — coloanele cu modelele noastre raman goale.')

    rows = []
    for r in harvest:
        their_car, their_years = parse_their_slug(r['modelLor'])
        # ce modele de-ale noastre ar fi atinse de intrarea asta
        hits = []
        for o in ours:
            s, _ = mcm.score(o['car'], o['years'], their_car, their_years)
            if s > 0:
                hits.append((s, o))
        hits.sort(key=lambda x: -x[0])

        lim = r.get('limitari')
        if not hits:
            rows.append({
                'Marca': their_car.split()[0].upper() if their_car else '',
                'Modelul nostru': '(niciun model potrivit in catalogul nostru)',
                'Anii nostri': '', 'Produse la noi': 0,
                'Modelul lor': their_car, 'Anii lor': their_years,
                'Are limitari?': 'DA' if lim else 'NU',
                'Textul lor (verbatim)': lim or '',
                'Tip': clasifica(lim) if lim else '',
                'Se aplica la noi?': '', 'Observatii': '',
                'Link': r['url'],
            })
            continue

        for s, o in hits:
            rows.append({
                'Marca': o.get('brand', ''),
                'Modelul nostru': o['car'],
                'Anii nostri': o['years'],
                'Produse la noi': o['products'],
                'Modelul lor': their_car,
                'Anii lor': their_years,
                'Are limitari?': 'DA' if lim else 'NU',
                'Textul lor (verbatim)': lim or '',
                'Tip': clasifica(lim) if lim else '',
                'Se aplica la noi?': '',
                'Observatii': '',
                'Link': r['url'],
            })

    cols = ['Marca', 'Modelul nostru', 'Anii nostri', 'Produse la noi',
            'Modelul lor', 'Anii lor', 'Are limitari?', 'Textul lor (verbatim)',
            'Tip', 'Se aplica la noi?', 'Observatii', 'Link']
    # sortare: intai cele cu limitari, apoi dupa cate produse ating
    rows.sort(key=lambda r: (r['Are limitari?'] != 'DA', -r['Produse la noi']))

    with open(CSV_OUT, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)

    cu = sum(1 for r in rows if r['Are limitari?'] == 'DA')
    prod = sum(r['Produse la noi'] for r in rows if r['Are limitari?'] == 'DA')
    print(f'Randuri: {len(rows)} | cu limitari: {cu} | produse de-ale noastre atinse: {prod}')
    print(f'Scris {CSV_OUT}')

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = Workbook(); ws = wb.active; ws.title = 'Limitari'
            ws.append(cols)
            for c in ws[1]:
                c.font = Font(bold=True, color='FFFFFF')
                c.fill = PatternFill('solid', fgColor='4472C4')
            for r in rows:
                ws.append([r[c] for c in cols])
            widths = {'A': 14, 'B': 26, 'C': 12, 'D': 13, 'E': 26, 'F': 12,
                      'G': 12, 'H': 70, 'I': 42, 'J': 17, 'K': 26, 'L': 60}
            for col, wd in widths.items():
                ws.column_dimensions[col].width = wd
            for row in ws.iter_rows(min_row=2):
                row[7].alignment = Alignment(wrap_text=True, vertical='top')
                row[8].alignment = Alignment(wrap_text=True, vertical='top')
            ws.freeze_panes = 'A2'
            wb.save(XLSX_OUT)
            print(f'Scris {XLSX_OUT}')
        except ImportError:
            print('openpyxl lipseste — doar CSV. (pip3 install openpyxl)')


if __name__ == '__main__':
    main()
